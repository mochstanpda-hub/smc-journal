"""
Záložka 🔁 REPLAY — analýza TradingView Replay Trading exportů.
Více XLSX souborů se sčítá do jednoho konzistentního přehledu.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os, json, math, uuid, shutil
from datetime import datetime

# ── Optional deps ─────────────────────────────────────────────────────────────
try:
    import openpyxl
    _HAS_XL = True
except ImportError:
    _HAS_XL = False

try:
    import matplotlib
    matplotlib.use('TkAgg')
    from matplotlib.figure import Figure
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    import matplotlib.patches as mpatches
    import matplotlib.ticker as mticker
    import matplotlib.gridspec as mgridspec
    from matplotlib.lines import Line2D
    _HAS_MPL = True
except ImportError:
    _HAS_MPL = False

# ── Colours ───────────────────────────────────────────────────────────────────
BG      = '#0f172a'
PANEL   = '#1e293b'
SURF    = '#293548'
SURF2   = '#1a2438'
TEXT    = '#e2e8f0'
SUB     = '#64748b'
ACCENT  = '#3b82f6'
GREEN   = '#22c55e'
RED     = '#ef4444'
ORANGE  = '#f59e0b'
BORDER  = '#334155'
MFIG    = '#1e293b'
MAXES   = '#233044'
MGRID   = '#2d3f55'
MTEXT   = '#e2e8f0'

FN   = ('Segoe UI', 10)
FNB  = ('Segoe UI', 10, 'bold')
FNS  = ('Segoe UI', 9)
FNXL = ('Segoe UI', 18, 'bold')
FNLG = ('Segoe UI', 12, 'bold')

# Stránky analýzy
PAGES = [
    ('overview', '  Přehled  '),
    ('equity',   '  Equity & Drawdown  '),
    ('trades',   '  Obchody  '),
    ('dist',     '  Rozdělení  '),
    ('time',     '  Čas  '),
    ('stats',    '  Statistiky  '),
]
WDAYS = ['Po', 'Út', 'St', 'Čt', 'Pá', 'So', 'Ne']

# ── Global state ──────────────────────────────────────────────────────────────
_APP_DIR = None


def _data_path():
    d = _APP_DIR or os.getcwd()
    return os.path.join(d, 'replay_data.json')


def _archive_dir():
    """Složka pro archivaci nahraných XLSX (uvnitř složky programu)."""
    d = os.path.join(_APP_DIR or os.getcwd(), 'replay_soubory')
    os.makedirs(d, exist_ok=True)
    return d


def _archive_xlsx(src, symbol):
    """Zkopíruje XLSX do archivu. Vrací cestu ke kopii, nebo None."""
    try:
        base = os.path.basename(src)
        stem, ext = os.path.splitext(base)
        stamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
        name  = f'{stamp}_{symbol}_{stem}{ext}'
        # Windows: ošetři nepovolené znaky
        for ch in '<>:"/\\|?*':
            name = name.replace(ch, '-')
        dst = os.path.join(_archive_dir(), name)
        if os.path.abspath(src) == os.path.abspath(dst):
            return dst
        shutil.copy2(src, dst)
        return dst
    except Exception:
        return None


def _load():
    p = _data_path()
    if os.path.exists(p):
        try:
            with open(p, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {'sessions': []}


def _save(data):
    with open(_data_path(), 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ── XLSX parser ───────────────────────────────────────────────────────────────

def _parse_xlsx(path):
    """Parse TradingView Replay XLSX. Returns {initial_capital, trades[]}."""
    wb = openpyxl.load_workbook(path, data_only=True)

    cap = 1000.0
    if 'Performance' in wb.sheetnames:
        for row in wb['Performance'].iter_rows(values_only=True):
            if row[0] == 'Initial capital' and row[1] is not None:
                try:
                    cap = float(row[1])
                except Exception:
                    pass
                break

    if 'Trades' not in wb.sheetnames:
        raise ValueError("Soubor neobsahuje list 'Trades'")

    rows = list(wb['Trades'].iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("Žádné obchody v souboru")

    td = {}
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        try:
            num  = int(row[0])
            typ  = str(row[1] or '')
            dt   = row[2]
            dt_s = dt.strftime('%Y-%m-%d %H:%M') if isinstance(dt, datetime) else str(dt or '')
            sig  = str(row[3] or '')
            price = float(row[4] or 0)
            pnl   = float(row[7] or 0)
            ppct  = float(row[8] or 0)
            fav   = float(row[10] or 0)
            adv   = float(row[12] or 0)
            bars  = int(row[16]) if row[16] is not None else 0

            if num not in td:
                td[num] = {}
            if 'entry' in typ.lower():
                td[num].update({
                    'entry_time': dt_s,
                    'entry_price': price,
                    'type': 'long' if 'long' in typ.lower() else 'short',
                })
            else:
                td[num].update({
                    'exit_time': dt_s,
                    'exit_price': price,
                    'pnl': pnl,
                    'pnl_pct': ppct,
                    'fav_excursion': fav,
                    'adv_excursion': adv,
                    'bars': bars,
                    'signal': sig,
                })
        except Exception:
            continue

    trades = []
    for num in sorted(td.keys()):
        t = {
            'trade_num': num, 'type': 'short',
            'entry_time': '', 'exit_time': '',
            'entry_price': 0.0, 'exit_price': 0.0,
            'pnl': 0.0, 'pnl_pct': 0.0,
            'fav_excursion': 0.0, 'adv_excursion': 0.0,
            'bars': 0, 'signal': '',
        }
        t.update(td[num])
        trades.append(t)

    return {'initial_capital': cap, 'trades': trades}


# ── Statistics ────────────────────────────────────────────────────────────────

def _active_trades(data):
    trades = []
    for s in data.get('sessions', []):
        if not s.get('enabled', True):
            continue
        for t in s.get('trades', []):
            t2 = dict(t)
            t2['_sid']   = s['id']
            t2['_sym']   = s.get('symbol', '')
            trades.append(t2)
    trades.sort(key=lambda t: t.get('exit_time', '') or '')
    return trades


def _compute(trades):
    if not trades:
        return None

    pnls = [t['pnl'] for t in trades]
    winners = [p for p in pnls if p > 0]
    losers  = [p for p in pnls if p < 0]
    bes     = [p for p in pnls if p == 0]
    total   = len(pnls)
    tot_pnl = sum(pnls)
    gross_p = sum(winners)
    gross_l = abs(sum(losers))

    # Equity curve
    equity = [0.0]
    for p in pnls:
        equity.append(equity[-1] + p)

    # Max drawdown
    peak   = 0.0
    max_dd = 0.0
    for v in equity:
        peak = max(peak, v)
        max_dd = max(max_dd, peak - v)

    pf = gross_p / gross_l if gross_l > 0 else (float('inf') if gross_p > 0 else 0.0)

    if len(pnls) > 1:
        mean = tot_pnl / total
        var  = sum((p - mean) ** 2 for p in pnls) / (total - 1)
        std  = math.sqrt(var) if var > 0 else 1e-9
        sharpe = (mean / std) * math.sqrt(total)
    else:
        sharpe = 0.0

    pcts = [t.get('pnl_pct', 0) for t in trades]

    # Série drawdownu (underwater)
    dd_series = []
    _pk = 0.0
    for v in equity:
        _pk = max(_pk, v)
        dd_series.append(v - _pk)

    # Nejdelší série výher / proher
    cur_w = cur_l = max_cw = max_cl = 0
    for p in pnls:
        if p > 0:
            cur_w += 1; cur_l = 0
        elif p < 0:
            cur_l += 1; cur_w = 0
        else:
            cur_w = cur_l = 0
        max_cw = max(max_cw, cur_w)
        max_cl = max(max_cl, cur_l)

    # Long / Short rozpad
    def _side(ts):
        if not ts:
            return {'n': 0, 'pnl': 0.0, 'wr': 0.0, 'avg': 0.0}
        ps = [t['pnl'] for t in ts]
        return {'n': len(ps), 'pnl': sum(ps),
                'wr': len([p for p in ps if p > 0]) / len(ps) * 100,
                'avg': sum(ps) / len(ps)}
    st_long  = _side([t for t in trades if t.get('type') == 'long'])
    st_short = _side([t for t in trades if t.get('type') == 'short'])

    # Sortino (downside deviation)
    downside = [p for p in pnls if p < 0]
    if downside and total > 1:
        dstd = math.sqrt(sum(p * p for p in downside) / total)
        sortino = (tot_pnl / total) / dstd * math.sqrt(total) if dstd > 0 else 0.0
    else:
        sortino = 0.0

    recovery = (tot_pnl / max_dd) if max_dd > 0 else (float('inf') if tot_pnl > 0 else 0.0)

    # Časové agregace podle času výstupu
    by_hour, by_wday, by_month = {}, {}, {}
    for t in trades:
        try:
            d = datetime.strptime((t.get('exit_time') or '')[:16], '%Y-%m-%d %H:%M')
        except Exception:
            continue
        by_hour.setdefault(d.hour, []).append(t['pnl'])
        by_wday.setdefault(d.weekday(), []).append(t['pnl'])
        by_month.setdefault(d.strftime('%Y-%m-%d'), []).append(t['pnl'])

    # Rozpad podle důvodu výstupu
    by_signal = {}
    for t in trades:
        by_signal.setdefault((t.get('signal') or '—').strip() or '—', []).append(t['pnl'])

    return {
        'total': total,
        'tot_pnl': tot_pnl,
        'gross_p': gross_p,
        'gross_l': gross_l,
        'winners': len(winners),
        'losers': len(losers),
        'bes': len(bes),
        'win_rate': len(winners) / total * 100,
        'pf': pf,
        'epayoff': tot_pnl / total,
        'avg_w': gross_p / len(winners) if winners else 0,
        'avg_l': gross_l / len(losers) if losers else 0,
        'largest_w': max(pnls),
        'largest_l': min(pnls),
        'avg_bars': sum(t.get('bars', 0) for t in trades) / total,
        'max_dd': max_dd,
        'max_dd_pct': (max_dd / max(equity) * 100) if max(equity) > 0 else 0.0,
        'sharpe': sharpe,
        'sortino': sortino,
        'recovery': recovery,
        'equity': equity,
        'dd_series': dd_series,
        'pnls': pnls,
        'pcts': pcts,
        'max_cw': max_cw,
        'max_cl': max_cl,
        'long': st_long,
        'short': st_short,
        'by_hour': by_hour,
        'by_wday': by_wday,
        'by_month': by_month,
        'by_signal': by_signal,
        'mfe': [t.get('fav_excursion', 0) for t in trades],
        'mae': [t.get('adv_excursion', 0) for t in trades],
        'bars_list': [t.get('bars', 0) for t in trades],
        'ratio_wl': (gross_p / len(winners)) / (gross_l / len(losers))
                    if winners and losers else 0.0,
    }


# ── Main UI class ─────────────────────────────────────────────────────────────

class ReplayUI:
    def __init__(self, parent):
        self._data    = _load()
        self._fig     = None
        self._canvas  = None
        self._sv      = {}
        self._kpis    = {}
        try:
            self._build(parent)
        except Exception as _e:
            import traceback as _tb
            tk.Label(parent,
                     text=f'❌ _build() selhalo:\n{_e}\n\n{_tb.format_exc()}',
                     bg='#0f172a', fg='#ef4444', font=('Segoe UI', 9),
                     justify='left', wraplength=900).pack(expand=True, padx=20)
            return
        try:
            self._refresh()
        except Exception as _e:
            import traceback as _tb
            tk.Label(parent,
                     text=f'❌ _refresh() selhalo:\n{_e}\n\n{_tb.format_exc()}',
                     bg='#0f172a', fg='#ef4444', font=('Segoe UI', 9),
                     justify='left', wraplength=900).pack(expand=True, padx=20)

    # ── Layout ────────────────────────────────────────────────────────────────

    def _build(self, outer):
        # outer může být ttk.Frame (nezná -bg) → vlastní tk kontejner
        parent = tk.Frame(outer, bg=BG)
        parent.pack(fill='both', expand=True)

        # ── Toolbar ──────────────────────────────────────────────────────────
        tb = tk.Frame(parent, bg=SURF2, height=44)
        tb.pack(fill='x', side='top')
        tb.pack_propagate(False)

        tk.Label(tb, text='🔁  REPLAY TRADING', bg=SURF2, fg=TEXT,
                 font=FNB).pack(side='left', padx=16, pady=10)

        tk.Button(tb, text='📂  Načíst XLSX', bg=ACCENT, fg='white',
                  relief='flat', font=FNB, padx=14, pady=4,
                  cursor='hand2', activebackground='#2563eb',
                  command=self._add_session).pack(side='right', padx=10, pady=6)

        # ── Main split ───────────────────────────────────────────────────────
        split = tk.Frame(parent, bg=BG)
        split.pack(fill='both', expand=True)

        # LEFT — session list
        left = tk.Frame(split, bg=PANEL, width=230)
        left.pack(side='left', fill='y')
        left.pack_propagate(False)
        self._left = left
        self._build_sessions_panel()

        # RIGHT — charts + KPIs
        right = tk.Frame(split, bg=BG)
        right.pack(side='left', fill='both', expand=True)
        self._build_right(right)

    def _build_sessions_panel(self):
        hdr = tk.Frame(self._left, bg=SURF2, height=36)
        hdr.pack(fill='x')
        hdr.pack_propagate(False)
        tk.Label(hdr, text='Relace', bg=SURF2, fg=SUB, font=FNS
                 ).pack(side='left', padx=12, pady=8)
        tk.Button(hdr, text='📁', bg=SURF2, fg=SUB, relief='flat', bd=0,
                  font=FNS, cursor='hand2', activebackground=SURF,
                  activeforeground=TEXT, command=self._open_archive
                  ).pack(side='right', padx=8)

        # Scrollable list
        self._sess_canvas = tk.Canvas(self._left, bg=PANEL, highlightthickness=0)
        sb = tk.Scrollbar(self._left, orient='vertical',
                          command=self._sess_canvas.yview)
        self._sess_canvas.configure(yscrollcommand=sb.set)
        sb.pack(side='right', fill='y')
        self._sess_canvas.pack(fill='both', expand=True)

        self._sess_frame = tk.Frame(self._sess_canvas, bg=PANEL)
        self._sess_win   = self._sess_canvas.create_window(
            (0, 0), window=self._sess_frame, anchor='nw')
        self._sess_frame.bind('<Configure>', self._on_sess_configure)
        self._sess_canvas.bind('<Configure>',
            lambda e: self._sess_canvas.itemconfig(self._sess_win, width=e.width))

    def _open_archive(self):
        d = _archive_dir()
        try:
            os.startfile(d)
        except Exception:
            messagebox.showinfo('Archiv XLSX', f'Nahrané soubory se ukládají do:\n{d}')

    def _on_sess_configure(self, _evt):
        self._sess_canvas.configure(
            scrollregion=self._sess_canvas.bbox('all'))

    def _rebuild_sessions_list(self):
        for w in self._sess_frame.winfo_children():
            w.destroy()

        sessions = self._data.get('sessions', [])
        if not sessions:
            tk.Label(self._sess_frame,
                     text='Žádné relace.\nNačti XLSX soubor.',
                     bg=PANEL, fg=SUB, font=FNS,
                     wraplength=200, justify='center').pack(pady=24)
            return

        for s in sessions:
            sid  = s['id']
            sym  = s.get('symbol', '?')
            dt   = s.get('uploaded', '')[:10]
            cnt  = len(s.get('trades', []))
            pnl  = sum(t['pnl'] for t in s.get('trades', []))
            col  = GREEN if pnl >= 0 else RED
            sign = '+' if pnl >= 0 else ''

            card = tk.Frame(self._sess_frame, bg=SURF, pady=8, padx=10)
            card.pack(fill='x', padx=6, pady=4)

            top = tk.Frame(card, bg=SURF); top.pack(fill='x')

            if sid not in self._sv:
                self._sv[sid] = tk.BooleanVar(value=s.get('enabled', True))

            tk.Checkbutton(
                top, variable=self._sv[sid], bg=SURF,
                activebackground=SURF, selectcolor=ACCENT,
                command=lambda s2=s, v=self._sv[sid]: self._toggle(s2, v),
            ).pack(side='left')
            tk.Label(top, text=f'{sym}', bg=SURF, fg=TEXT,
                     font=FNB).pack(side='left')

            # Delete button
            tk.Button(top, text='✕', bg=SURF, fg=SUB, relief='flat',
                      font=('Segoe UI', 8), padx=4, cursor='hand2',
                      activebackground=RED, activeforeground='white',
                      command=lambda s2=s: self._delete(s2)
                      ).pack(side='right')

            row2 = tk.Frame(card, bg=SURF); row2.pack(fill='x', pady=(2, 0))
            tk.Label(row2, text=dt, bg=SURF, fg=SUB, font=FNS
                     ).pack(side='left')
            tk.Label(row2, text=f'{sign}${pnl:.2f}',
                     bg=SURF, fg=col, font=FNS).pack(side='right')

            row3 = tk.Frame(card, bg=SURF); row3.pack(fill='x')
            tk.Label(row3, text=f'{cnt} obchodů', bg=SURF, fg=SUB,
                     font=FNS).pack(side='left')
            fname = s.get('filename', '')
            if len(fname) > 22:
                fname = '…' + fname[-20:]
            tk.Label(row3, text=fname, bg=SURF, fg=SURF2, font=FNS
                     ).pack(side='right')

    def _build_right(self, parent):
        # ── KPI row ──────────────────────────────────────────────────────────
        kpi_row = tk.Frame(parent, bg=PANEL, height=80)
        kpi_row.pack(fill='x', padx=0, pady=0)
        kpi_row.pack_propagate(False)

        self._kpi_frame = kpi_row
        self._kpi_vars  = {}
        self._kpi_lbls  = {}

        kpi_defs = [
            ('tot_pnl',  'Total PnL',        '$0.00'),
            ('max_dd',   'Max Drawdown',      '$0.00'),
            ('win_rate', 'Profitable trades', '0 %'),
            ('pf',       'Profit factor',     '0.000'),
            ('sharpe',   'Sharpe ratio',      '0.000'),
            ('epayoff',  'Expected payoff',   '$0.00'),
        ]
        for i, (key, lbl, default) in enumerate(kpi_defs):
            sep_col = BORDER if i > 0 else PANEL
            tk.Frame(kpi_row, bg=sep_col, width=1).pack(side='left', fill='y', padx=0)
            cell = tk.Frame(kpi_row, bg=PANEL, padx=18)
            cell.pack(side='left', fill='y')
            tk.Label(cell, text=lbl, bg=PANEL, fg=SUB, font=FNS).pack(pady=(12, 0))
            var = tk.StringVar(value=default)
            self._kpi_vars[key] = var
            vlbl = tk.Label(cell, textvariable=var, bg=PANEL, fg=TEXT,
                            font=FNLG)
            vlbl.pack()
            self._kpi_lbls[key] = vlbl

        # ── Přepínač stránek ─────────────────────────────────────────────────
        navbar = tk.Frame(parent, bg=SURF2, height=38)
        navbar.pack(fill='x')
        navbar.pack_propagate(False)

        self._page = 'overview'
        self._page_btns = {}
        for key, lbl in PAGES:
            b = tk.Button(navbar, text=lbl, bg=SURF2, fg=SUB,
                          relief='flat', font=FNS, padx=14, pady=6,
                          cursor='hand2', bd=0,
                          activebackground=SURF, activeforeground=TEXT,
                          command=lambda k=key: self._set_page(k))
            b.pack(side='left', padx=1, pady=4)
            self._page_btns[key] = b

        # ── Chart canvas ─────────────────────────────────────────────────────
        chart_frame = tk.Frame(parent, bg=BG)
        chart_frame.pack(fill='both', expand=True, padx=6, pady=6)
        self._chart_frame = chart_frame

        if _HAS_MPL:
            self._fig = Figure(facecolor=MFIG)
            self._canvas = FigureCanvasTkAgg(self._fig, master=chart_frame)
            self._canvas.get_tk_widget().pack(fill='both', expand=True)
        else:
            tk.Label(chart_frame,
                     text='Chybí matplotlib. pip install matplotlib',
                     bg=BG, fg=RED, font=FN).pack(expand=True)

        # Textová tabulka statistik (schovaná, střídá se s grafem)
        self._stats_frame = tk.Frame(chart_frame, bg=BG)

        self._paint_page_btns()

    def _set_page(self, key):
        self._page = key
        self._paint_page_btns()
        self._refresh_charts()

    def _paint_page_btns(self):
        for k, b in self._page_btns.items():
            if k == self._page:
                b.config(bg=ACCENT, fg='white', font=FNB)
            else:
                b.config(bg=SURF2, fg=SUB, font=FNS)

    # ── Actions ───────────────────────────────────────────────────────────────

    def _add_session(self):
        if not _HAS_XL:
            messagebox.showerror('Chybí openpyxl',
                                 'Nainstaluj: pip install openpyxl')
            return
        paths = filedialog.askopenfilenames(
            title='Vyber TradingView Replay XLSX',
            filetypes=[('Excel', '*.xlsx'), ('Všechny', '*.*')],
        )
        if not paths:
            return
        added = 0
        for path in paths:
            try:
                parsed = _parse_xlsx(path)
                fname  = os.path.basename(path)
                # Detect symbol from filename (e.g. Replay_Trading_SKILLING_US100_...)
                sym = 'N/A'
                parts = fname.replace('.xlsx', '').split('_')
                # Try to find a capitalised ticker-like token
                for p in reversed(parts):
                    if p.isupper() and 2 <= len(p) <= 8 and p not in ('TRADING', 'REPLAY', 'SKILLING'):
                        sym = p
                        break

                archived = _archive_xlsx(path, sym)

                session = {
                    'id':       str(uuid.uuid4()),
                    'filename': fname,
                    'symbol':   sym,
                    'uploaded': datetime.now().strftime('%Y-%m-%d'),
                    'enabled':  True,
                    'initial_capital': parsed['initial_capital'],
                    'trades':   parsed['trades'],
                    'archive':  archived or '',
                    'source':   path,
                }
                self._data.setdefault('sessions', []).append(session)
                added += 1
            except Exception as e:
                messagebox.showerror('Chyba', f'{os.path.basename(path)}:\n{e}')

        if added:
            _save(self._data)
            self._refresh()

    def _delete(self, session):
        if not messagebox.askyesno('Smazat relaci',
                                   f'Smazat relaci "{session.get("filename","")}"?\n'
                                   'Všechna data relace budou odstraněna.'):
            return
        sid = session['id']
        self._data['sessions'] = [
            s for s in self._data['sessions'] if s['id'] != sid
        ]
        self._sv.pop(sid, None)
        _save(self._data)
        self._refresh()

    def _toggle(self, session, var):
        session['enabled'] = var.get()
        _save(self._data)
        self._refresh()

    # ── Refresh ───────────────────────────────────────────────────────────────

    def _refresh(self):
        self._rebuild_sessions_list()
        self._refresh_charts()

    def _refresh_charts(self):
        trades = _active_trades(self._data)
        stats  = _compute(trades)
        self._update_kpis(stats)
        if _HAS_MPL and self._fig is not None:
            self._draw_charts(stats, trades)

    def _update_kpis(self, st):
        if st is None:
            for k, v in self._kpi_vars.items():
                v.set('—')
            return
        pnl_sign = '+' if st['tot_pnl'] >= 0 else ''
        self._kpi_vars['tot_pnl'].set(f"{pnl_sign}${st['tot_pnl']:.2f}")
        self._kpi_vars['max_dd'].set(f"${st['max_dd']:.2f}")
        self._kpi_vars['win_rate'].set(
            f"{st['win_rate']:.1f}%  {st['winners']}/{st['total']}")
        pf_val = st['pf']
        self._kpi_vars['pf'].set('∞' if math.isinf(pf_val) else f"{pf_val:.3f}")
        self._kpi_vars['sharpe'].set(f"{st['sharpe']:.3f}")
        ep_sign = '+' if st['epayoff'] >= 0 else ''
        self._kpi_vars['epayoff'].set(f"{ep_sign}${st['epayoff']:.2f}")

        # Obarvení hodnot podle znaménka
        self._kpi_lbls['tot_pnl'].config(
            fg=GREEN if st['tot_pnl'] >= 0 else RED)
        self._kpi_lbls['max_dd'].config(fg=RED if st['max_dd'] > 0 else TEXT)
        self._kpi_lbls['pf'].config(
            fg=GREEN if st['pf'] >= 1 else RED)
        self._kpi_lbls['epayoff'].config(
            fg=GREEN if st['epayoff'] >= 0 else RED)

    # ── Charts ────────────────────────────────────────────────────────────────

    def _ax_style(self, ax, title='', xlabel='', ylabel=''):
        ax.set_facecolor(MAXES)
        ax.tick_params(colors=MTEXT, labelsize=8)
        ax.spines[:].set_color(MGRID)
        ax.yaxis.label.set_color(MTEXT)
        ax.xaxis.label.set_color(MTEXT)
        ax.grid(color=MGRID, linewidth=0.5, linestyle='--', alpha=0.6)
        if title:
            ax.set_title(title, color=MTEXT, fontsize=9, pad=6)
        if xlabel:
            ax.set_xlabel(xlabel, fontsize=8)
        if ylabel:
            ax.set_ylabel(ylabel, fontsize=8)

    def _show_canvas(self):
        """Zobrazí matplotlib plátno, schová textovou tabulku."""
        self._stats_frame.pack_forget()
        w = self._canvas.get_tk_widget()
        if not w.winfo_ismapped():
            w.pack(fill='both', expand=True)

    def _show_stats_table(self):
        self._canvas.get_tk_widget().pack_forget()
        if not self._stats_frame.winfo_ismapped():
            self._stats_frame.pack(fill='both', expand=True)

    def _msg(self, text, color=None):
        """Vykreslí zprávu přes celé plátno."""
        self._show_canvas()
        self._fig.clf()
        self._fig.patch.set_facecolor(MFIG)
        ax = self._fig.add_subplot(111)
        ax.set_facecolor(MAXES)
        ax.text(0.5, 0.5, text, transform=ax.transAxes,
                ha='center', va='center', color=color or SUB,
                fontsize=10, linespacing=1.6, family='monospace'
                if color else 'sans-serif')
        ax.set_xticks([]); ax.set_yticks([])
        ax.spines[:].set_color(MGRID)
        self._canvas.draw()

    def _draw_charts(self, st, trades):
        if st is None:
            self._msg('Načti XLSX soubor pro zobrazení analýzy')
            return

        if self._page == 'stats':
            self._show_stats_table()
            try:
                self._build_stats_table(st)
            except Exception as e:
                import traceback as tb
                for w in self._stats_frame.winfo_children():
                    w.destroy()
                tk.Label(self._stats_frame, text=f'{e}\n\n{tb.format_exc()}',
                         bg=BG, fg=RED, font=FNS, justify='left',
                         wraplength=900).pack(padx=20, pady=20)
            return

        self._show_canvas()
        self._fig.clf()
        self._fig.patch.set_facecolor(MFIG)
        try:
            {
                'overview': self._page_overview,
                'equity':   self._page_equity,
                'trades':   self._page_trades,
                'dist':     self._page_dist,
                'time':     self._page_time,
            }[self._page](st, trades)
            self._canvas.draw()
        except Exception as e:
            import traceback as tb
            self._msg(f'Chyba vykreslení ({self._page}):\n\n'
                      f'{tb.format_exc()}', color=RED)

    # ── Stránky ───────────────────────────────────────────────────────────────

    def _grid(self, rows, cols, **kw):
        gs = mgridspec
        opts = dict(figure=self._fig, hspace=0.45, wspace=0.30,
                    left=0.07, right=0.97, top=0.92, bottom=0.10)
        opts.update(kw)
        return gs.GridSpec(rows, cols, **opts)

    def _page_overview(self, st, trades):
        grid = self._grid(2, 3, height_ratios=[2, 1])
        self._draw_equity(self._fig.add_subplot(grid[0, :]), st, trades)
        self._draw_profit_structure(self._fig.add_subplot(grid[1, 0]), st)
        self._draw_donut(self._fig.add_subplot(grid[1, 1]), st)
        self._draw_histogram(self._fig.add_subplot(grid[1, 2]), st)

    def _page_equity(self, st, trades):
        grid = self._grid(2, 1, height_ratios=[2.2, 1], hspace=0.28, bottom=0.12)
        ax_eq = self._fig.add_subplot(grid[0, 0])
        ax_dd = self._fig.add_subplot(grid[1, 0], sharex=ax_eq)
        self._draw_equity(ax_eq, st, trades, big=True)
        self._draw_underwater(ax_dd, st, trades)
        ax_eq.tick_params(labelbottom=False)   # popisky jen na spodní ose
        ax_eq.set_xlabel('')
        ax_dd.set_xlabel('Čas výstupu', fontsize=8)

    def _page_trades(self, st, trades):
        grid = self._grid(2, 2)
        self._draw_pertrade(self._fig.add_subplot(grid[0, 0]), st)
        self._draw_mae_mfe(self._fig.add_subplot(grid[0, 1]), st)
        self._draw_duration(self._fig.add_subplot(grid[1, 0]), st)
        self._draw_long_short(self._fig.add_subplot(grid[1, 1]), st)

    def _page_dist(self, st, trades):
        grid = self._grid(2, 2)
        self._draw_histogram(self._fig.add_subplot(grid[0, 0]), st)
        self._draw_pnl_hist(self._fig.add_subplot(grid[0, 1]), st)
        self._draw_streaks(self._fig.add_subplot(grid[1, 0]), st)
        self._draw_signals(self._fig.add_subplot(grid[1, 1]), st)

    def _page_time(self, st, trades):
        grid = self._grid(2, 2, height_ratios=[1, 1])
        self._draw_by_hour(self._fig.add_subplot(grid[0, :]), st)
        self._draw_by_wday(self._fig.add_subplot(grid[1, 0]), st)
        self._draw_by_day(self._fig.add_subplot(grid[1, 1]), st)

    def _draw_equity(self, ax, st, trades, big=False):
        self._ax_style(ax, title='Equity křivka — kumulativní P&L')
        pnls   = st['pnls']
        equity = st['equity']
        n      = len(pnls)
        xs     = list(range(1, n + 1))

        colors = [GREEN if p >= 0 else RED for p in pnls]
        ax.bar(xs, pnls, color=colors, alpha=0.6, width=0.7, zorder=2)

        eq = equity[1:]
        ax.plot(xs, eq, color='#38bdf8', linewidth=2,
                zorder=3, marker='o', markersize=3.5, markerfacecolor='#38bdf8')
        ax.fill_between(xs, 0, eq, color='#38bdf8', alpha=0.10, zorder=1)

        if big:
            # Běžící maximum (peak) — vizualizace drawdownu
            peaks, pk = [], float('-inf')
            for v in eq:
                pk = max(pk, v)
                peaks.append(pk)
            ax.plot(xs, peaks, color=SUB, linewidth=1, linestyle='--',
                    zorder=2, label='Peak')
            ax.fill_between(xs, eq, peaks, color=RED, alpha=0.14, zorder=1)

        # Zero line
        ax.axhline(0, color=MGRID, linewidth=0.8, linestyle='-')

        # Annotate final value
        final = equity[-1]
        col   = GREEN if final >= 0 else RED
        ax.annotate(f'${final:.2f}',
                    xy=(n, final), xytext=(n + 0.5, final),
                    color=col, fontsize=8, va='center',
                    arrowprops=dict(arrowstyle='-', color=col, lw=0.5))

        # X-axis labels: use exit time for context
        step = max(1, n // 10)
        ticks = list(range(0, n, step)) + [n - 1]
        ticks = sorted(set(ticks))
        labels = []
        for i in ticks:
            t = trades[i] if i < len(trades) else None
            labels.append(t.get('exit_time', '')[:10] if t else '')
        ax.set_xticks([i + 1 for i in ticks])
        ax.set_xticklabels(labels, rotation=30, ha='right', fontsize=7)
        ax.set_xlim(0, n + 2)
        ax.set_ylabel('P&L (USD)', fontsize=8)

        # Legend
        pw = mpatches.Patch(color=GREEN, alpha=0.7, label='WIN')
        lw = mpatches.Patch(color=RED,   alpha=0.7, label='LOSS')
        el = ax.lines[0] if ax.lines else None
        handles = [pw, lw]
        if el:
            handles.append(Line2D([0], [0], color='#38bdf8', lw=2, label='Kumulativní P&L'))
        ax.legend(handles=handles, loc='upper left', fontsize=7,
                  facecolor=MFIG, edgecolor=MGRID, labelcolor=MTEXT)

    def _draw_profit_structure(self, ax, st):
        self._ax_style(ax, title='Profit struktura')
        labels = ['Gross\nProfit', 'Gross\nLoss', 'Net\nPnL']
        values = [st['gross_p'], -st['gross_l'], st['tot_pnl']]
        colors = [GREEN, RED, GREEN if st['tot_pnl'] >= 0 else RED]
        bars = ax.bar(labels, values, color=colors, alpha=0.8, width=0.5)
        ax.axhline(0, color=MGRID, linewidth=0.8)
        for bar, val in zip(bars, values):
            ax.text(bar.get_x() + bar.get_width() / 2,
                    val + (abs(val) * 0.03 if val >= 0 else -abs(val) * 0.05),
                    f'${val:.1f}', ha='center', va='bottom' if val >= 0 else 'top',
                    color=MTEXT, fontsize=8)
        ax.set_ylabel('USD', fontsize=8)
        ax.yaxis.set_major_formatter(
            mticker.FuncFormatter(lambda x, _: f'${x:.0f}'))

    def _draw_donut(self, ax, st):
        ax.set_facecolor(MAXES)
        ax.set_title('Distribuce obchodů', color=MTEXT, fontsize=9, pad=6)

        w  = st['winners']
        l  = st['losers']
        be = st['bes']
        sizes  = [x for x in [w, l, be] if x > 0]
        labels = [x for x, v in [('Winners', w), ('Losers', l), ('Breakevens', be)] if v > 0]
        colors = [x for x, v in [(GREEN, w), (RED, l), (ORANGE, be)] if v > 0]

        if not sizes:
            ax.text(0.5, 0.5, 'Žádná data', transform=ax.transAxes,
                    ha='center', va='center', color=SUB)
            return

        wedges, _ = ax.pie(
            sizes, labels=None, colors=colors,
            wedgeprops={'width': 0.55, 'edgecolor': MFIG, 'linewidth': 1.5},
            startangle=90,
        )
        ax.text(0, 0, f'{st["total"]}\nobchodů',
                ha='center', va='center', color=MTEXT,
                fontsize=9, fontweight='bold', linespacing=1.5)

        legend_labels = [
            f'{lbl}  {v}  ({v/st["total"]*100:.1f}%)'
            for lbl, v in zip(['Winners', 'Losers', 'BEs'], [w, l, be])
            if v > 0
        ]
        patches = [mpatches.Patch(color=c, label=lb)
                   for c, lb in zip(colors, legend_labels)]
        ax.legend(handles=patches, loc='lower center',
                  bbox_to_anchor=(0.5, -0.28),
                  fontsize=7, facecolor=MFIG,
                  edgecolor=MGRID, labelcolor=MTEXT, ncol=1)

    def _draw_histogram(self, ax, st):
        self._ax_style(ax, title='Distribuce výnosů (%)')
        pcts = st['pcts']
        if not pcts:
            return

        wins_pct  = [p for p in pcts if p > 0]
        loses_pct = [p for p in pcts if p < 0]

        bins = 20
        if loses_pct:
            ax.hist(loses_pct, bins=bins, color=RED,   alpha=0.75, label='Losers')
        if wins_pct:
            ax.hist(wins_pct,  bins=bins, color=GREEN, alpha=0.75, label='Winners')

        # Average lines
        if wins_pct:
            avg_w = sum(wins_pct) / len(wins_pct)
            ax.axvline(avg_w, color=GREEN, linestyle='--', linewidth=1.2,
                       label=f'Avg W {avg_w:.2f}%')
        if loses_pct:
            avg_l = sum(loses_pct) / len(loses_pct)
            ax.axvline(avg_l, color=RED, linestyle='--', linewidth=1.2,
                       label=f'Avg L {avg_l:.2f}%')

        ax.axvline(0, color=MTEXT, linewidth=0.6)
        ax.set_xlabel('%', fontsize=8)
        ax.set_ylabel('Počet', fontsize=8)
        ax.legend(fontsize=7, facecolor=MFIG, edgecolor=MGRID,
                  labelcolor=MTEXT, loc='upper right')

    # ── Nové grafy ────────────────────────────────────────────────────────────

    def _draw_underwater(self, ax, st, trades):
        self._ax_style(ax, title='Drawdown (underwater)')
        dd = st['dd_series'][1:]
        xs = list(range(1, len(dd) + 1))
        ax.fill_between(xs, dd, 0, color=RED, alpha=0.45, zorder=2)
        ax.plot(xs, dd, color=RED, linewidth=1.2, zorder=3)
        ax.axhline(0, color=MGRID, linewidth=0.8)
        if dd:
            worst = min(dd)
            wi    = dd.index(worst) + 1
            ax.annotate(f'max DD ${abs(worst):.2f}',
                        xy=(wi, worst), xytext=(wi, worst),
                        color=RED, fontsize=8, ha='center', va='top')
        ax.set_ylabel('USD pod peakem', fontsize=8)
        ax.set_xlabel('Obchod #', fontsize=8)

    def _draw_pertrade(self, ax, st):
        self._ax_style(ax, title='P&L jednotlivých obchodů')
        pnls = st['pnls']
        xs   = list(range(1, len(pnls) + 1))
        ax.bar(xs, pnls, color=[GREEN if p >= 0 else RED for p in pnls],
               alpha=0.85, width=0.7)
        ax.axhline(0, color=MGRID, linewidth=0.8)
        ep = st['epayoff']
        ax.axhline(ep, color=ORANGE, linestyle='--', linewidth=1.1,
                   label=f'Průměr ${ep:.2f}')
        ax.set_xlabel('Obchod #', fontsize=8)
        ax.set_ylabel('USD', fontsize=8)
        ax.legend(fontsize=7, facecolor=MFIG, edgecolor=MGRID,
                  labelcolor=MTEXT, loc='best')

    def _draw_mae_mfe(self, ax, st):
        self._ax_style(ax, title='MAE / MFE — jak daleko šel obchod')
        mae  = [abs(v) for v in st['mae']]
        mfe  = [abs(v) for v in st['mfe']]
        pnls = st['pnls']
        if not mae:
            return
        cols = [GREEN if p >= 0 else RED for p in pnls]
        ax.scatter(mae, mfe, c=cols, s=42, alpha=0.85,
                   edgecolors=MFIG, linewidths=0.8, zorder=3)
        lim = max(max(mae), max(mfe)) * 1.1 or 1
        ax.plot([0, lim], [0, lim], color=SUB, linestyle='--',
                linewidth=0.9, zorder=2, label='MAE = MFE')
        ax.set_xlim(0, lim); ax.set_ylim(0, lim)
        ax.set_xlabel('MAE — max proti pozici', fontsize=8)
        ax.set_ylabel('MFE — max pro pozici', fontsize=8)
        h = [mpatches.Patch(color=GREEN, label='WIN'),
             mpatches.Patch(color=RED,   label='LOSS')]
        ax.legend(handles=h, fontsize=7, facecolor=MFIG,
                  edgecolor=MGRID, labelcolor=MTEXT, loc='upper left')

    def _draw_duration(self, ax, st):
        self._ax_style(ax, title='Délka obchodu (počet svíček)')
        bl = [b for b in st['bars_list'] if b]
        if not bl:
            ax.text(0.5, 0.5, 'Žádná data o délce', transform=ax.transAxes,
                    ha='center', va='center', color=SUB, fontsize=9)
            return
        bins = min(20, max(5, len(set(bl))))
        ax.hist(bl, bins=bins, color=ACCENT, alpha=0.8)
        avg = sum(bl) / len(bl)
        ax.axvline(avg, color=ORANGE, linestyle='--', linewidth=1.2,
                   label=f'Průměr {avg:.1f}')
        ax.set_xlabel('Svíčky', fontsize=8)
        ax.set_ylabel('Počet obchodů', fontsize=8)
        ax.legend(fontsize=7, facecolor=MFIG, edgecolor=MGRID,
                  labelcolor=MTEXT, loc='upper right')

    def _draw_long_short(self, ax, st):
        self._ax_style(ax, title='Long vs Short')
        L, S = st['long'], st['short']
        labels = ['Long', 'Short']
        vals   = [L['pnl'], S['pnl']]
        cols   = [GREEN if v >= 0 else RED for v in vals]
        bars   = ax.bar(labels, vals, color=cols, alpha=0.85, width=0.45)
        ax.axhline(0, color=MGRID, linewidth=0.8)
        for b, v, sd in zip(bars, vals, [L, S]):
            ax.text(b.get_x() + b.get_width() / 2,
                    v + (abs(v) * 0.04 if v >= 0 else -abs(v) * 0.04),
                    f'${v:.1f}\n{sd["n"]} obch. · {sd["wr"]:.0f}% WR',
                    ha='center', va='bottom' if v >= 0 else 'top',
                    color=MTEXT, fontsize=8, linespacing=1.4)
        ax.set_ylabel('USD', fontsize=8)
        lo, hi = ax.get_ylim()
        ax.set_ylim(lo * 1.3 if lo < 0 else lo, hi * 1.3 if hi > 0 else hi)

    def _draw_pnl_hist(self, ax, st):
        self._ax_style(ax, title='Rozdělení P&L (USD)')
        pnls = st['pnls']
        if not pnls:
            return
        w = [p for p in pnls if p > 0]
        l = [p for p in pnls if p < 0]
        if l:
            ax.hist(l, bins=15, color=RED, alpha=0.75, label='Losers')
        if w:
            ax.hist(w, bins=15, color=GREEN, alpha=0.75, label='Winners')
        ax.axvline(0, color=MTEXT, linewidth=0.6)
        ax.set_xlabel('USD', fontsize=8)
        ax.set_ylabel('Počet', fontsize=8)
        ax.legend(fontsize=7, facecolor=MFIG, edgecolor=MGRID,
                  labelcolor=MTEXT, loc='upper right')

    def _draw_streaks(self, ax, st):
        self._ax_style(ax, title='Série za sebou')
        pnls = st['pnls']
        runs, cur, sign = [], 0, 0
        for p in pnls:
            s = 1 if p > 0 else (-1 if p < 0 else 0)
            if s == sign and s != 0:
                cur += 1
            else:
                if sign != 0:
                    runs.append(sign * cur)
                sign, cur = s, (1 if s != 0 else 0)
        if sign != 0:
            runs.append(sign * cur)
        if not runs:
            ax.text(0.5, 0.5, 'Žádná data', transform=ax.transAxes,
                    ha='center', va='center', color=SUB)
            return
        xs = list(range(1, len(runs) + 1))
        ax.bar(xs, runs, color=[GREEN if r > 0 else RED for r in runs],
               alpha=0.85, width=0.7)
        ax.axhline(0, color=MGRID, linewidth=0.8)
        ax.set_xlabel('Pořadí série', fontsize=8)
        ax.set_ylabel('Délka (+W / −L)', fontsize=8)
        ax.yaxis.set_major_formatter(
            mticker.FuncFormatter(lambda v, _: f'{abs(int(v))}'))
        ax.text(0.02, 0.95,
                f'Max výher v řadě: {st["max_cw"]}\nMax proher v řadě: {st["max_cl"]}',
                transform=ax.transAxes, va='top', ha='left',
                color=MTEXT, fontsize=8, linespacing=1.5)

    def _draw_signals(self, ax, st):
        self._ax_style(ax, title='Důvod výstupu')
        bs = st['by_signal']
        if not bs:
            ax.text(0.5, 0.5, 'Žádná data', transform=ax.transAxes,
                    ha='center', va='center', color=SUB)
            return
        items = sorted(bs.items(), key=lambda kv: sum(kv[1]))
        names = [k if len(k) <= 22 else k[:20] + '…' for k, _ in items]
        vals  = [sum(v) for _, v in items]
        cnts  = [len(v) for _, v in items]
        ypos  = list(range(len(names)))
        ax.barh(ypos, vals, color=[GREEN if v >= 0 else RED for v in vals],
                alpha=0.85, height=0.6)
        ax.set_yticks(ypos)
        ax.set_yticklabels(names, fontsize=7)
        ax.axvline(0, color=MGRID, linewidth=0.8)
        for y, v, c in zip(ypos, vals, cnts):
            ax.text(v, y, f'  ${v:.1f} ({c}×)  ',
                    va='center', ha='left' if v >= 0 else 'right',
                    color=MTEXT, fontsize=7)
        ax.set_xlabel('USD', fontsize=8)
        lo, hi = ax.get_xlim()
        ax.set_xlim(lo * 1.35 if lo < 0 else lo, hi * 1.35 if hi > 0 else hi)

    def _bucket_bar(self, ax, buckets, labels, title, xlabel):
        self._ax_style(ax, title=title)
        if not buckets:
            ax.text(0.5, 0.5, 'Žádná data', transform=ax.transAxes,
                    ha='center', va='center', color=SUB)
            return
        vals = [sum(buckets[k]) for k in buckets]
        cnts = [len(buckets[k]) for k in buckets]
        xs   = list(range(len(labels)))
        ax.bar(xs, vals, color=[GREEN if v >= 0 else RED for v in vals],
               alpha=0.85, width=0.6)
        ax.axhline(0, color=MGRID, linewidth=0.8)
        ax.set_xticks(xs)
        ax.set_xticklabels(labels, fontsize=7, rotation=0)
        for x, v, c in zip(xs, vals, cnts):
            ax.text(x, v, f'{c}×', ha='center',
                    va='bottom' if v >= 0 else 'top',
                    color=SUB, fontsize=7)
        ax.set_xlabel(xlabel, fontsize=8)
        ax.set_ylabel('USD', fontsize=8)
        lo, hi = ax.get_ylim()
        ax.set_ylim(lo * 1.25 if lo < 0 else lo, hi * 1.25 if hi > 0 else hi)

    def _draw_by_hour(self, ax, st):
        bh = st['by_hour']
        ks = sorted(bh.keys())
        self._bucket_bar(ax, {k: bh[k] for k in ks},
                         [f'{k:02d}:00' for k in ks],
                         'P&L podle hodiny výstupu', 'Hodina')

    def _draw_by_wday(self, ax, st):
        bw = st['by_wday']
        ks = sorted(bw.keys())
        self._bucket_bar(ax, {k: bw[k] for k in ks},
                         [WDAYS[k] for k in ks],
                         'P&L podle dne v týdnu', 'Den')

    def _draw_by_day(self, ax, st):
        bm = st['by_month']
        ks = sorted(bm.keys())
        self._bucket_bar(ax, {k: bm[k] for k in ks},
                         [k[5:] for k in ks],
                         'P&L podle dne', 'Datum')

    # ── Tabulka statistik ─────────────────────────────────────────────────────

    def _build_stats_table(self, st):
        for w in self._stats_frame.winfo_children():
            w.destroy()

        def money(v):
            return ('+' if v > 0 else '') + f'${v:,.2f}'

        def num(v, d=3):
            return '∞' if math.isinf(v) else f'{v:.{d}f}'

        L, S = st['long'], st['short']
        groups = [
            ('Výkonnost', [
                ('Čistý zisk (Net P&L)',   money(st['tot_pnl']), st['tot_pnl']),
                ('Hrubý zisk',             money(st['gross_p']), 1),
                ('Hrubá ztráta',           money(-st['gross_l']), -1),
                ('Profit factor',          num(st['pf']), st['pf'] - 1),
                ('Expected payoff',        money(st['epayoff']), st['epayoff']),
                ('Recovery factor',        num(st['recovery'], 2), st['recovery'] - 1),
            ]),
            ('Riziko', [
                ('Max drawdown',           f"${st['max_dd']:,.2f}", -1),
                ('Max drawdown %',         f"{st['max_dd_pct']:.2f} %", -1),
                ('Sharpe ratio',           num(st['sharpe']), st['sharpe']),
                ('Sortino ratio',          num(st['sortino']), st['sortino']),
            ]),
            ('Obchody', [
                ('Celkem obchodů',         str(st['total']), 0),
                ('Ziskových',              f"{st['winners']}  ({st['win_rate']:.1f} %)", 1),
                ('Ztrátových',             f"{st['losers']}  ({st['losers']/st['total']*100:.1f} %)", -1),
                ('Breakeven',              str(st['bes']), 0),
                ('Průměrný zisk',          money(st['avg_w']), 1),
                ('Průměrná ztráta',        money(-st['avg_l']), -1),
                ('Poměr avg W / avg L',    num(st['ratio_wl'], 2), st['ratio_wl'] - 1),
                ('Největší zisk',          money(st['largest_w']), 1),
                ('Největší ztráta',        money(st['largest_l']), -1),
                ('Max výher v řadě',       str(st['max_cw']), 1),
                ('Max proher v řadě',      str(st['max_cl']), -1),
                ('Průměrná délka (svíčky)', f"{st['avg_bars']:.1f}", 0),
            ]),
            ('Long vs Short', [
                ('Long — počet / P&L',     f"{L['n']}  ·  {money(L['pnl'])}", L['pnl']),
                ('Long — úspěšnost',       f"{L['wr']:.1f} %", L['wr'] - 50),
                ('Short — počet / P&L',    f"{S['n']}  ·  {money(S['pnl'])}", S['pnl']),
                ('Short — úspěšnost',      f"{S['wr']:.1f} %", S['wr'] - 50),
            ]),
        ]

        wrap = tk.Frame(self._stats_frame, bg=BG)
        wrap.pack(fill='both', expand=True, padx=10, pady=10)

        for i, (title, rows) in enumerate(groups):
            col = tk.Frame(wrap, bg=PANEL)
            col.grid(row=i // 2, column=i % 2, sticky='nsew', padx=6, pady=6)

            tk.Label(col, text=title, bg=SURF2, fg=TEXT, font=FNB,
                     anchor='w', padx=12, pady=7).pack(fill='x')

            for j, (lbl, val, sign) in enumerate(rows):
                bgc = PANEL if j % 2 == 0 else SURF2
                r = tk.Frame(col, bg=bgc)
                r.pack(fill='x')
                tk.Label(r, text=lbl, bg=bgc, fg=SUB, font=FNS,
                         anchor='w', padx=12, pady=5).pack(side='left')
                fg = TEXT if sign == 0 else (GREEN if sign > 0 else RED)
                tk.Label(r, text=val, bg=bgc, fg=fg, font=FNB,
                         anchor='e', padx=12, pady=5).pack(side='right')

        wrap.grid_columnconfigure(0, weight=1)
        wrap.grid_columnconfigure(1, weight=1)
        for r in range((len(groups) + 1) // 2):
            wrap.grid_rowconfigure(r, weight=1)


# ── Public entry point ────────────────────────────────────────────────────────

def setup_replay_tab(parent, app_dir):
    """Inicializuje záložku Replay. Volat z BACKTESTING.py."""
    global _APP_DIR
    _APP_DIR = app_dir

    missing = []
    if not _HAS_XL:
        missing.append('openpyxl')
    if not _HAS_MPL:
        missing.append('matplotlib')

    if missing:
        import subprocess as _sp
        frame = tk.Frame(parent, bg=BG)
        frame.pack(expand=True)
        tk.Label(frame,
                 text=f'❌ Chybí knihovny: {", ".join(missing)}',
                 bg=BG, fg=RED, font=('Segoe UI', 13, 'bold')).pack(pady=(40, 8))
        tk.Label(frame,
                 text='Klikni na tlačítko níže pro instalaci, pak restartuj program.',
                 bg=BG, fg=TEXT, font=('Segoe UI', 10)).pack(pady=(0, 20))

        def _install():
            import sys as _s
            # Hledej Python.exe (ne exe programu)
            import shutil as _sh
            py = _sh.which('python') or _sh.which('python3')
            if not py:
                from tkinter import messagebox as _mb
                _mb.showerror('Chyba', 'Python nenalezen v PATH.\nNainstaluj ručně:\npip install ' + ' '.join(missing))
                return
            btn.config(state='disabled', text='⏳ Instaluji…')
            try:
                _sp.check_call([py, '-m', 'pip', 'install', '--quiet'] + missing,
                               creationflags=0x08000000)  # CREATE_NO_WINDOW
                from tkinter import messagebox as _mb
                _mb.showinfo('Hotovo', 'Instalace dokončena.\nRestartuj program.')
            except Exception as _e:
                from tkinter import messagebox as _mb
                _mb.showerror('Chyba', f'Instalace selhala:\n{_e}\n\nZkus ručně:\npip install {" ".join(missing)}')
            btn.config(state='normal', text='📦  Nainstalovat a restartovat')

        btn = tk.Button(frame, text='📦  Nainstalovat a restartovat',
                        bg=ACCENT, fg='white', relief='flat',
                        font=('Segoe UI', 11, 'bold'), padx=20, pady=10,
                        cursor='hand2', command=_install)
        btn.pack()
        return

    try:
        ReplayUI(parent)
    except Exception as _e:
        import traceback as _tb
        tk.Label(parent,
                 text=f'❌ Chyba inicializace Replay:\n{_e}\n\n{_tb.format_exc()}',
                 bg='#0f172a', fg='#ef4444', font=('Segoe UI', 9),
                 justify='left', wraplength=900).pack(expand=True, padx=20)
